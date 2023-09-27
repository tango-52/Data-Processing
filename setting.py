import yaml
import pandas as pd
from openpyxl import load_workbook


# yaml文件读取
def read_config_file(config_file):
    with open(config_file, 'r', encoding='utf-8') as config_file:
        config_data = yaml.load(config_file, Loader=yaml.FullLoader)
    return config_data


# 读取Excel文件
def read_excel_file(file_path):
    df = pd.read_excel(file_path)
    return df


# 写入Excel, row、column同Excel坐标值保持一致
def rewrite_excel(file_path, modified_data, row=12, column=1):
    wb = load_workbook(file_path)
    ws = wb.active
    cell = ws.cell(row, column)
    cell.value = modified_data
    wb.save(file_path)
