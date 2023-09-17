import pandas as pd
import yaml
from openpyxl import load_workbook

# 读取TDOC数据格式处理文件
file_path = '数据文件/TDOC数据格式.xlsx'
df = pd.read_excel(file_path)

# 读取 YAML 配置文件
with open('tdoc_data_config.yaml', 'r', encoding='utf-8') as config_file:
    config_data = yaml.load(config_file, Loader=yaml.FullLoader)

positions = config_data['positions']
lengths = config_data['lengths']

replace_values = [
    '23',  # buyerInvoiceNo 长度15
    '23',  # blNo  长度20
    '23',  # blDate  长度8
    '23'   # invoiceNo  长度7
]

# 创建原始数据的副本
df_copy = df.copy()

# 获取原始数据
origin_data = df.iloc[9, 0]

# 使用openpyxl来写入数据，并保留原始的公式或函数
wb = load_workbook('数据文件/TDOC数据格式.xlsx')
ws = wb.active


def process_data(position, lengths, replace_value, modified_data):
    """
    替换原始数据中的值
    """
    if len(replace_value) <= lengths:
        replace_fix_value = replace_value.ljust(lengths)
    else:
        replace_fix_value = replace_value[:lengths]

    modified_data = modified_data[:position-1] + replace_fix_value + modified_data[position-1+lengths:]   # 将原有的子串值进行替换

    return modified_data


for position, length, replace_value in zip(positions, lengths, replace_values):
    origin_data = process_data(position, length, replace_value, origin_data)

# 将修改后的数据写入到Excel文件中的对应单元格
cell = ws.cell(row=12, column=1)
cell.value = origin_data

# 输出最终结果
print(origin_data)

# 保存工作簿
wb.save(file_path)
