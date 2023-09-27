import yaml
import pandas as pd
from openpyxl import load_workbook


# yaml文件读取
def read_config_file(config_file):
    with open(config_file, 'r', encoding='utf-8') as config_file:
        config_data = yaml.load(config_file, Loader=yaml.FullLoader)
    return config_data


# 读取Excel文件
def read_excel_file(file_path):
    df = pd.read_excel(file_path)
    return df


# 写入TDOCExcel
def rewrite_excel(file_path, modified_data, row=12, column=1):
    wb = load_workbook(file_path)
    ws = wb.active
    cell = ws.cell(row, column)
    cell.value = modified_data
    wb.save(file_path)


# 判断替换value长度 进行截取或者补齐内容
def re_value(replace_value, lengths):
    if isinstance(replace_value, str):
        if len(replace_value) <= lengths:
            replace_fix_value = replace_value.ljust(lengths)
        else:
            replace_fix_value = replace_value[:lengths]
        return replace_fix_value
    else:
        # 处理非字符串情况的逻辑
        str_replace_value =  str(replace_value)  # 将其他类型转换为字符串
        if len(str_replace_value) <= lengths:
            replace_fix_value = replace_value.ljust(lengths)
        else:
            replace_fix_value = replace_value[:lengths]
        return replace_fix_value

# 重新拼接数据
def concat_value(modified_data, position, replace_fix_value, lengths):
    modified_data = modified_data[:position - 1] + replace_fix_value \
                    + modified_data[position - 1 + lengths:]
    return modified_data


def conf_data(config_data):
    positions = config_data['positions']
    lengths = config_data['lengths']
    replace_values = config_data['replace_values']
    data_locations = config_data['data_locations']
    return positions, lengths, replace_values, data_locations


def process_data(position, lengths, replace_value, modified_data):
    """
    替换原始数据中的值
    """
    replace_fix_value = re_value(replace_value, lengths)
    modified_data = concat_value(modified_data, position, replace_fix_value, lengths)  # 将原有的子串值进行替换

    return modified_data
